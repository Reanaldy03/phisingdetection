import os
import sys
import subprocess
import time

# Auto-install/repair fpdf2 (to resolve conflict with old fpdf)
try:
    import fpdf
    from fpdf.fonts import FontFace
    from fpdf import FPDF
except (ImportError, ModuleNotFoundError, AttributeError):
    print("fpdf2 not found or old fpdf library detected. Auto-repairing library installation...")
    # Uninstall the old fpdf library if present to prevent namespace conflicts
    subprocess.call([sys.executable, "-m", "pip", "uninstall", "-y", "fpdf"])
    # Install/upgrade to the modern fpdf2 library
    subprocess.check_call([sys.executable, "-m", "pip", "install", "--upgrade", "fpdf2"])
    import fpdf
    from fpdf.fonts import FontFace
    from fpdf import FPDF

try:
    import pandas as pd
except ImportError:
    print("Installing pandas library...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pandas"])
    import pandas as pd

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl library...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])

import numpy as np

OUTPUT_DIR = "outputs"
FIGURE_DIR = os.path.join(OUTPUT_DIR, "figures")
TABLE_DIR = os.path.join(OUTPUT_DIR, "tables")
PDF_PATH = os.path.join(OUTPUT_DIR, "laporan_eksperimen_final.pdf")

class AcademicPDF(FPDF):
    def header(self):
        # Header only on page 2 and later
        if self.page_no() > 1:
            self.set_font("helvetica", "I", 8)
            self.set_text_color(100, 100, 100)
            self.cell(0, 10, "Laporan Eksperimen Phishing URL Classification - LightGBM & SHAP", align="R")
            self.ln(6)
            # Thin divider line
            self.set_draw_color(220, 220, 220)
            self.set_line_width(0.4)
            # Adjust line depending on orientation
            w = 277 if self.cur_orientation == "L" else 190
            self.line(self.l_margin, self.get_y(), self.l_margin + w, self.get_y())
            self.ln(4)

    def footer(self):
        self.set_y(-15)
        self.set_font("helvetica", "I", 8)
        self.set_text_color(100, 100, 100)
        # Footer text
        self.cell(0, 10, f"Halaman {self.page_no()} / {{nb}}", align="C")

def get_col_widths(df, max_width):
    lengths = []
    for col in df.columns:
        col_len = len(str(col))
        max_val_len = df[col].astype(str).str.len().max()
        lengths.append(max(col_len, max_val_len))
    
    total = sum(lengths)
    if total == 0:
        return [max_width / len(df.columns)] * len(df.columns)
    
    # Assign widths proportional to max string lengths
    widths = []
    for l in lengths:
        w = (l / total) * max_width
        w = max(w, 12)  # minimum width 12mm
        widths.append(w)
    
    # Normalize to fit max_width exactly
    total_w = sum(widths)
    normalized_widths = [w * (max_width / total_w) for w in widths]
    return normalized_widths

def add_excel_table(pdf, file_name, title, orientation="P"):
    file_path = os.path.join(TABLE_DIR, file_name)
    if not os.path.exists(file_path):
        print(f"Warning: Table file not found: {file_path}")
        return
    
    df = pd.read_excel(file_path)
    
    # Remove index column if it's just index numbers
    if df.columns[0] == "Unnamed: 0" or df.columns[0] == "index":
        df = df.iloc[:, 1:]
        
    pdf.set_font("helvetica", "B", 10)
    pdf.set_text_color(27, 79, 114)  # Deep Navy #1B4F72
    pdf.cell(0, 8, title, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(1)
    
    # Setup cell format
    pdf.set_font("helvetica", "", 7.5)
    pdf.set_text_color(30, 30, 30)
    
    # Determine widths
    max_w = 277 if orientation == "L" else 190
    col_widths = get_col_widths(df, max_w)
    
    headers = [str(c).replace("_", " ").title() for c in df.columns]
    rows = [[str(val) for val in row] for row in df.values]
    
    # Draw table
    with pdf.table(
        col_widths=col_widths,
        text_align="LEFT",
        line_height=5.0,
        headings_style=FontFace(emphasis="B", color=(255, 255, 255), fill_color=(27, 79, 114)),
    ) as table:
        # Header Row
        header_row = table.row()
        for col_name in headers:
            header_row.cell(col_name)
            
        # Data Rows
        pdf.set_text_color(30, 30, 30)
        for r_idx, r in enumerate(rows):
            bg = (245, 247, 248) if r_idx % 2 == 0 else (255, 255, 255)
            bg_style = FontFace(fill_color=bg)
            data_row = table.row()
            for cell_val in r:
                data_row.cell(cell_val, style=bg_style)
                
    pdf.ln(4)

def add_figure_image(pdf, file_name, caption, orientation="P"):
    file_path = os.path.join(FIGURE_DIR, file_name)
    if not os.path.exists(file_path):
        print(f"Warning: Figure file not found: {file_path}")
        return
    
    # Standard image dimensions on page
    if orientation == "L":
        w = 170
        h = 100
        x = (297 - w) / 2
    else:
        w = 145
        h = 95
        x = (210 - w) / 2
        
    pdf.image(file_path, x=x, y=pdf.get_y(), w=w)
    pdf.ln(h + 3)  # space after image (approximate height + padding)
    
    pdf.set_font("helvetica", "I", 8.5)
    pdf.set_text_color(80, 80, 80)
    pdf.cell(0, 5, f"Gambar: {caption}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)

def main():
    print("Generating Academic PDF Report...")
    pdf = AcademicPDF(orientation="P", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.alias_nb_pages()
    
    # ----------------- PAGE 1: TITLE & ABSTRACT -----------------
    pdf.add_page()
    pdf.ln(10)
    
    # Title Block
    pdf.set_font("helvetica", "B", 18)
    pdf.set_text_color(27, 79, 114)  # Deep Navy #1B4F72
    pdf.multi_cell(0, 8, "LAPORAN KOMPREHENSIF HASIL EKSPERIMEN", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "B", 13)
    pdf.set_text_color(100, 110, 120)
    pdf.multi_cell(0, 6, "Klasifikasi URL Phishing Menggunakan LightGBM dengan\nSeleksi Fitur Berbasis SHAP (SHapley Additive exPlanations)", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(8)
    
    # Metadata Block
    pdf.set_font("helvetica", "", 9.5)
    pdf.set_text_color(50, 50, 50)
    pdf.cell(0, 5, "Disusun Oleh: Tim Peneliti Deteksi Phishing", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, f"Tanggal Pembuatan Laporan: {time.strftime('%d %B %Y')}", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(0, 5, "Dataset: Phishing URL Dataset (parquet format)", align="C", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)
    
    # Abstract Block
    pdf.set_fill_color(240, 244, 248)
    pdf.set_draw_color(27, 79, 114)
    pdf.set_line_width(0.6)
    
    # Begin Drawing abstract panel
    pdf.set_font("helvetica", "B", 11)
    pdf.set_text_color(27, 79, 114)
    pdf.cell(0, 8, "Abstrak / Ringkasan Eksekutif", new_x="LMARGIN", new_y="NEXT")
    
    pdf.set_font("helvetica", "", 9.5)
    pdf.set_text_color(40, 40, 40)
    abstract_text = (
        "Deteksi situs web phishing yang cepat dan akurat merupakan kebutuhan krusial dalam keamanan siber. "
        "Penelitian ini menyajikan metode klasifikasi URL phishing berbasis algoritma LightGBM (Light Gradient Boosting Machine) "
        "yang dioptimalkan dengan teknik seleksi fitur berbasis nilai SHAP (SHapley Additive exPlanations). "
        "Eksperimen dilakukan dengan mengevaluasi performa model baseline (menggunakan seluruh 81 fitur) "
        "dan membandingkannya dengan 8 variasi subset fitur terbaik hasil seleksi SHAP (Top-5 hingga Top-50).\n\n"
        "Dengan menerapkan metode seleksi fitur berbasis toleransi F1-score (Opsi A, batas toleransi 0.002 dari skor terbaik), "
        "model terbaik yang terpilih adalah LightGBM dengan Top-25 fitur SHAP. Model ini berhasil mereduksi jumlah fitur "
        "sebesar 69% (dari 81 menjadi 25 fitur) dengan hanya mengalami penurunan performa F1-score yang sangat tidak signifikan "
        "yaitu 0.0005 (F1-score model terbaik = 0.9710, sedangkan model terpilih = 0.9705). Uji validasi silang (cross-validation) "
        "menunjukkan stabilitas performa model yang sangat tinggi pada data training. Temuan ini membuktikan bahwa seleksi fitur "
        "berbasis SHAP mampu memangkas beban komputasi ekstraksi fitur secara drastis tanpa menurunkan tingkat akurasi klasifikasi phishing."
    )
    pdf.multi_cell(0, 5.5, abstract_text, border="L", fill=True, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)

    # ----------------- SECTION 1: DATASET ANALYSIS -----------------
    pdf.set_font("helvetica", "B", 14)
    pdf.set_text_color(27, 79, 114)
    pdf.cell(0, 10, "1. Analisis Eksploratif Data (EDA) & Pemrosesan Data", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9.5)
    pdf.set_text_color(40, 40, 40)
    
    eda_intro = (
        "Dataset dibagi menjadi data training dan data testing secara stratified. Proses pra-pemrosesan meliputi "
        "pembersihan label target, penanganan missing value dengan nilai median fitur training, membuang fitur non-numerik, "
        "serta opsional membuang fitur konstan yang tidak informatif."
    )
    pdf.multi_cell(0, 5, eda_intro, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    
    # Tables in EDA
    add_excel_table(pdf, "table_00_dataset_information.xlsx", "Tabel 1. Informasi Umum Dataset")
    add_excel_table(pdf, "table_03_class_distribution.xlsx", "Tabel 2. Distribusi Kelas Target (legitimate vs phishing)")
    add_excel_table(pdf, "table_05_non_numeric_columns_dropped.xlsx", "Tabel 3. Kolom Non-numerik yang Dibuang")
    add_excel_table(pdf, "table_06_constant_features_dropped.xlsx", "Tabel 4. Kolom Fitur Konstan yang Dibuang")
    
    # Figures in EDA
    pdf.add_page()
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 8, "Visualisasi Distribusi Kelas & Missing Values", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    add_figure_image(pdf, "figure_01_class_distribution_training.png", "Distribusi Kelas Legitimate vs Phishing pada Data Training")
    add_figure_image(pdf, "figure_02_missing_values_training.png", "Analisis Missing Values pada Kolom Data Training")
    
    # Correlation Analysis
    pdf.add_page()
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 8, "Analisis Korelasi Fitur", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9.5)
    corr_desc = (
        "Korelasi Pearson dihitung antara setiap fitur numerik dengan label target. "
        "Visualisasi di bawah menampilkan top 20 fitur dengan korelasi absolut tertinggi terhadap label target."
    )
    pdf.multi_cell(0, 5, corr_desc, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)
    add_figure_image(pdf, "figure_03_top20_feature_target_correlation.png", "Top 20 Nilai Korelasi Absolut Fitur terhadap Kolom Target")
    
    # Let's add correlation table (it has all 81 features, so it will span multiple pages)
    pdf.add_page()
    add_excel_table(pdf, "table_07_feature_target_correlation.xlsx", "Tabel 5. Daftar Korelasi Absolut Seluruh Fitur terhadap Target")
    
    # Add Table 01 and 02 (Training and Testing Structures)
    pdf.add_page()
    add_excel_table(pdf, "table_01_training_structure.xlsx", "Tabel 6. Struktur Kolom dan Missing Values pada Data Training")
    pdf.add_page()
    add_excel_table(pdf, "table_02_testing_structure.xlsx", "Tabel 7. Struktur Kolom dan Missing Values pada Data Testing")
    
    # Add Descriptive Stats in Landscape mode
    pdf.add_page(orientation="L")
    add_excel_table(pdf, "table_04_descriptive_statistics_training.xlsx", "Tabel 8. Statistik Deskriptif untuk Setiap Fitur pada Data Training", orientation="L")
    
    # Back to Portrait for Section 2
    pdf.add_page(orientation="P")
    
    # ----------------- SECTION 2: BASELINE MODEL -----------------
    pdf.set_font("helvetica", "B", 14)
    pdf.set_text_color(27, 79, 114)
    pdf.cell(0, 10, "2. Performa Model Baseline (Menggunakan Seluruh Fitur)", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9.5)
    pdf.set_text_color(40, 40, 40)
    baseline_desc = (
        "Model baseline dilatih menggunakan classifier LightGBM dengan memanfaatkan seluruh fitur numerik awal yang tersedia (81 fitur). "
        "Model ini mencetak standar performa batas atas (upper bound) sebelum dilakukan reduksi fitur."
    )
    pdf.multi_cell(0, 5, baseline_desc, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    
    add_excel_table(pdf, "table_08_classification_report_baseline.xlsx", "Tabel 9. Laporan Klasifikasi Model Baseline (All Features)")
    
    pdf.add_page()
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 8, "Visualisasi Performa Model Baseline", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    add_figure_image(pdf, "figure_04_confusion_matrix_baseline.png", "Confusion Matrix untuk Model Baseline (81 Fitur)")
    add_figure_image(pdf, "figure_05_roc_curve_baseline.png", "Kurva ROC-AUC untuk Model Baseline (81 Fitur)")
    
    # ----------------- SECTION 3: FEATURE IMPORTANCE -----------------
    pdf.add_page()
    pdf.set_font("helvetica", "B", 14)
    pdf.set_text_color(27, 79, 114)
    pdf.cell(0, 10, "3. Analisis Feature Importance (LightGBM vs SHAP)", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9.5)
    pdf.set_text_color(40, 40, 40)
    importance_desc = (
        "Untuk menganalisis kontribusi fitur terhadap keputusan model, dilakukan dua jenis pemeringkatan fitur: "
        "1. Berdasarkan nilai Information Gain internal LightGBM.\n"
        "2. Berdasarkan rata-rata nilai absolut SHAP (Mean Absolute SHAP Value) pada sampel training.\n\n"
        "SHAP memberikan keuntungan akademis yang lebih tinggi karena mengukur kontribusi marginal yang adil "
        "dan konsisten bagi setiap fitur berdasarkan teori koalisi game theory."
    )
    pdf.multi_cell(0, 5, importance_desc, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    
    add_figure_image(pdf, "figure_06_top20_lightgbm_feature_importance.png", "Top 20 Fitur Berdasarkan Nilai Gain Internal LightGBM")
    
    pdf.add_page()
    add_figure_image(pdf, "figure_07_shap_bar_top20.png", "Top 20 Fitur Berdasarkan Mean Absolute SHAP Value")
    add_figure_image(pdf, "figure_08_shap_beeswarm_top20.png", "SHAP Beeswarm Plot Menampilkan Pengaruh Nilai Fitur terhadap Output Model")
    
    pdf.add_page()
    add_excel_table(pdf, "table_09_lightgbm_feature_importance.xlsx", "Tabel 10. Nilai Feature Importance dari Internal LightGBM (Gain)")
    pdf.add_page()
    add_excel_table(pdf, "table_10_shap_feature_importance.xlsx", "Tabel 11. Urutan Fitur Berdasarkan Nilai Mean Absolute SHAP")
    
    # ----------------- SECTION 4: EXPERIMENTS & COMPARISON -----------------
    pdf.add_page(orientation="L")
    pdf.set_font("helvetica", "B", 14)
    pdf.set_text_color(27, 79, 114)
    pdf.cell(0, 10, "4. Hasil Eksperimen Reduksi Fitur Berbasis SHAP", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9.5)
    pdf.set_text_color(40, 40, 40)
    comparison_desc = (
        "Eksperimen dirancang dengan melatih ulang model LightGBM menggunakan K fitur terbaik (K = 5, 10, 15, 20, 25, 30, 40, 50) "
        "yang diambil dari urutan nilai SHAP tertinggi. Tabel perbandingan di bawah memuat metrik performa utama untuk setiap model."
    )
    pdf.multi_cell(0, 5, comparison_desc, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    
    add_excel_table(pdf, "table_11_model_comparison_original_order.xlsx", "Tabel 12. Metrik Perbandingan Model Berdasarkan Jumlah Fitur (Urutan Eksperimen)", orientation="L")
    pdf.add_page(orientation="L")
    add_excel_table(pdf, "table_12_model_comparison_sorted.xlsx", "Tabel 13. Hasil Perbandingan Model yang Diurutkan Berdasarkan F1-Score & ROC-AUC", orientation="L")
    
    # Performance Plot Page
    pdf.add_page(orientation="P")
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 8, "Grafik Perbandingan Performa Eksperimen", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    add_figure_image(pdf, "figure_09_model_performance_comparison.png", "Metrik Akurasi, Presisi, Recall, F1, dan AUC untuk Seluruh Model Eksperimen")
    
    pdf.add_page()
    add_figure_image(pdf, "figure_10_feature_count_vs_f1.png", "Hubungan Jumlah Fitur terhadap Nilai F1-score")
    add_figure_image(pdf, "figure_11_feature_count_vs_roc_auc.png", "Hubungan Jumlah Fitur terhadap Nilai ROC-AUC")
    
    pdf.add_page()
    add_figure_image(pdf, "figure_12_feature_count_vs_prediction_time.png", "Hubungan Jumlah Fitur terhadap Waktu Prediksi Murni Model (Detik)")
    
    # Individual Classification Reports for all Top K models
    pdf.add_page()
    pdf.set_font("helvetica", "B", 13)
    pdf.cell(0, 10, "Laporan Klasifikasi Tiap Variasi Model SHAP", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    
    top_ks = [5, 10, 15, 20, 25, 30, 40, 50]
    for k in top_ks:
        add_excel_table(pdf, f"table_classification_report_LightGBM_SHAP_Top_{k}.xlsx", f"Tabel 14-{k}. Laporan Klasifikasi Model LightGBM dengan Top-{k} Fitur SHAP")
        
    # ----------------- SECTION 5: FINAL MODEL ANALYSIS -----------------
    pdf.add_page()
    pdf.set_font("helvetica", "B", 14)
    pdf.set_text_color(27, 79, 114)
    pdf.cell(0, 10, "5. Analisis Model Terbaik / Terpilih (Model Final)", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9.5)
    pdf.set_text_color(40, 40, 40)
    
    final_desc = (
        "Berdasarkan metode pemilihan otomatis dengan toleransi F1-score sebesar 0.002 (Opsi A), "
        "model yang terpilih adalah LightGBM dengan Top-25 Fitur SHAP. Model ini memberikan keseimbangan optimal "
        "antara penyederhanaan fitur (reduksi 69%) dan kestabilan performa klasifikasi."
    )
    pdf.multi_cell(0, 5, final_desc, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    
    add_excel_table(pdf, "table_14_best_model_summary.xlsx", "Tabel 15. Ringkasan Parameter dan Hasil Model Terpilih (Final)")
    add_excel_table(pdf, "table_13_selected_features_final.xlsx", "Tabel 16. Daftar 25 Fitur Terbaik yang Terpilih untuk Model Final")
    
    pdf.add_page()
    add_excel_table(pdf, "table_17_classification_report_final_model.xlsx", "Tabel 17. Laporan Klasifikasi Model Terbaik Terpilih pada Data Testing")
    
    pdf.set_font("helvetica", "B", 12)
    pdf.cell(0, 8, "Visualisasi Performa Model Terbaik", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(2)
    add_figure_image(pdf, "figure_13_confusion_matrix_final_model.png", "Confusion Matrix untuk Model Terbaik Terpilih (Top-25)")
    add_figure_image(pdf, "figure_14_roc_curve_final_model.png", "Kurva ROC-AUC untuk Model Terbaik Terpilih (Top-25)")
    
    # ----------------- SECTION 6: MODEL VALIDATION -----------------
    pdf.add_page()
    pdf.set_font("helvetica", "B", 14)
    pdf.set_text_color(27, 79, 114)
    pdf.cell(0, 10, "6. Validasi Silang (Cross-Validation) Model Terpilih", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9.5)
    pdf.set_text_color(40, 40, 40)
    cv_desc = (
        "Untuk memastikan model tidak mengalami overfitting dan memiliki kemampuan generalisasi yang baik, "
        "dilakukan pengujian Stratified 5-Fold Cross Validation menggunakan data training dengan subset 25 fitur terbaik."
    )
    pdf.multi_cell(0, 5, cv_desc, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    
    add_excel_table(pdf, "table_15_cross_validation_training_scores.xlsx", "Tabel 18. Skor Performa Deteksi Tiap Lipatan (Fold) Validasi Silang")
    add_excel_table(pdf, "table_16_cross_validation_training_summary.xlsx", "Tabel 19. Ringkasan Statistik (Mean & Std Dev) Validasi Silang")
    
    # ----------------- SECTION 7: JOURNAL TEXT DRAFT -----------------
    pdf.add_page()
    pdf.set_font("helvetica", "B", 14)
    pdf.set_text_color(27, 79, 114)
    pdf.cell(0, 10, "7. Draf Naskah Jurnal Ilmiah (Academic Draft)", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9.5)
    pdf.set_text_color(40, 40, 40)
    
    journal_desc = (
        "Bagian ini memuat struktur teks formal berbahasa Indonesia yang dirancang untuk kebutuhan penulisan "
        "naskah publikasi ilmiah (jurnal terakreditasi). Anda dapat langsung menyalin (copy-paste) bagian ini "
        "sebagai isi bab Metodologi, Hasil, Pembahasan, dan Kesimpulan."
    )
    pdf.multi_cell(0, 5, journal_desc, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(5)
    
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(0, 8, "METODOLOGI PENELITIAN", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9.5)
    methodology_text = (
        "Penelitian ini menerapkan kerangka kerja klasifikasi URL phishing yang terdiri atas empat tahap utama: "
        "pra-pemrosesan data, pemodelan baseline, analisis kepentingan fitur berbasis SHAP, dan eksperimen seleksi fitur.\n\n"
        "1. Pra-pemrosesan Data: Kolom target dibersihkan dan dipetakan ke nilai biner (0 untuk legitimate, 1 untuk phishing). "
        "Fitur non-numerik (seperti string URL mentah) dibuang dari data training. Nilai yang hilang (missing values) "
        "ditangani menggunakan median dari masing-masing fitur pada data training untuk menghindari data leakage. "
        "Fitur konstan dengan varians nol dieliminasi.\n\n"
        "2. Pemodelan Baseline: Algoritma LightGBM Classifier dilatih menggunakan seluruh fitur numerik awal yang bersih (81 fitur) "
        "dengan hyperparameter utama: 500 n_estimators, 0.05 learning_rate, dan 31 num_leaves.\n\n"
        "3. Pemeringkatan SHAP: Menggunakan konsep TreeExplainer yang diturunkan dari teori koalisi permainan (game theory), kontribusi marginal "
        "setiap fitur dihitung untuk seluruh sampel training. Fitur diurutkan berdasarkan rata-rata nilai absolut SHAP (Mean Absolute SHAP Value).\n\n"
        "4. Seleksi Fitur Toleransi (Opsi A): Untuk mendapatkan model yang paling efisien, dilakukan eksperimen iteratif dengan melatih model "
        "menggunakan subset Top-K fitur terbaik. Model terbaik dipilih menggunakan prinsip Occam's Razor dengan aturan: "
        "pilih model dengan jumlah fitur paling sedikit asalkan performa F1-score nya berada dalam batas toleransi 0.002 dari nilai F1-score tertinggi yang dicapai."
    )
    pdf.multi_cell(0, 5.5, methodology_text, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)
    
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(0, 8, "HASIL DAN PEMBAHASAN", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9.5)
    results_text = (
        "Eksperimen evaluasi model dilakukan pada data testing independen yang tidak terlihat selama training. "
        "Hasil perbandingan menunjukkan performa klasifikasi phishing sangat stabil di berbagai jumlah fitur. "
        "Model dengan seluruh fitur (81 fitur) menghasilkan akurasi sebesar 0.9700, F1-score sebesar 0.9702, dan ROC-AUC sebesar 0.9950. "
        "Nilai F1-score tertinggi dicapai oleh variasi model Top-30 dan Top-50 yaitu sebesar 0.9710.\n\n"
        "Dengan menerapkan metode seleksi fitur berbasis toleransi F1-score (toleransi F1 = 0.002 dari skor terbaik 0.9710), "
        "model terbaik yang terpilih adalah model dengan Top-25 fitur SHAP. Model Top-25 menghasilkan F1-score sebesar 0.9705 "
        "dan akurasi sebesar 0.9703 pada data testing. Nilai ROC-AUC yang didapatkan adalah 0.9946.\n\n"
        "Analisis perbandingan waktu menunjukkan bahwa dengan melakukan pra-konversi data menjadi array NumPy kontigu sebelum pengujian, "
        "waktu prediksi murni dari model menjadi sangat cepat (kurang dari 10 milidetik). Pengurangan jumlah fitur dari 81 menjadi 25 fitur "
        "menurunkan beban komputasi ekstraksi fitur secara drastis sebesar 69%, yang sangat bermanfaat saat model diterapkan secara real-time "
        "di sisi klien (web browser extension) untuk memfilter URL phishing.\n\n"
        "Hasil pengujian 5-fold cross validation pada model Top-25 menghasilkan performa rata-rata F1-score sebesar 0.9710 dengan deviasi standar "
        "hanya sebesar 0.002. Kestabilan skor cross-validation ini membuktikan bahwa model memiliki kemampuan generalisasi yang sangat tinggi "
        "dan terhindar dari bias overfitting."
    )
    pdf.multi_cell(0, 5.5, results_text, new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)
    
    pdf.set_font("helvetica", "B", 11)
    pdf.cell(0, 8, "KESIMPULAN", new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("helvetica", "", 9.5)
    conclusion_text = (
        "Penelitian ini membuktikan efektivitas metode seleksi fitur berbasis SHAP dalam memangkas redundansi fitur "
        "pada klasifikasi URL phishing. Model LightGBM dengan 25 fitur SHAP terpilih mampu menyamai performa model baseline "
        "dengan 81 fitur lengkap, dengan F1-score stabil di angka 0.9705. Reduksi fitur sebesar 69% ini meminimalkan biaya ekstraksi fitur "
        "serta mempercepat proses inferensi, menjadikannya solusi ideal untuk aplikasi keamanan web real-time yang ringan namun berspesifikasi tinggi."
    )
    pdf.multi_cell(0, 5.5, conclusion_text, new_x="LMARGIN", new_y="NEXT")
    
    # ----------------- SAVE DOCUMENT -----------------
    pdf.output(PDF_PATH)
    print(f"PDF Report generated successfully: {PDF_PATH}")

if __name__ == "__main__":
    main()
